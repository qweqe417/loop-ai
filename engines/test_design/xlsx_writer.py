"""xlsx 写入器 —— 从 TestDesignBundle 生成人读 Excel。

视图A 的唯一输出文件 —— 3 个 Sheet:
  Sheet 1「测试用例」—— 主表，列因 scope 不同
  Sheet 2「覆盖矩阵」—— 需求 vs 用例
  Sheet 3「未决问题」—— 如有

规则: 不写 API 路径，用业务描述。
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import TestDesignBundle, Scope

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_data(ws, start_row: int, end_row: int, ncols: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN


def _col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_xlsx(bundle: TestDesignBundle, filepath: str | Path) -> Path:
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    _write_sheet_test_cases(wb, bundle)
    _write_sheet_coverage(wb, bundle)
    _write_sheet_open_questions(wb, bundle)

    # 覆盖率小结放第一个 sheet 末尾
    _write_summary_row(wb, bundle)

    wb.save(str(filepath))
    logger.info("xlsx written to %s (%d cases)", filepath, len(bundle.test_cases))
    return filepath


def _write_sheet_test_cases(wb: Workbook, bundle: TestDesignBundle) -> None:
    scope = bundle.scope
    ws = wb.active
    ws.title = "测试用例"

    if scope == Scope.BACKEND:
        headers = ["编号", "标题", "优先级", "前置条件", "测试步骤", "预期结果", "数据校验点"]
        widths = [16, 28, 8, 26, 40, 36, 36]
    elif scope == Scope.FRONTEND:
        headers = ["编号", "标题", "优先级", "前置条件", "操作步骤", "预期页面表现", "UI 校验点"]
        widths = [16, 28, 8, 26, 40, 36, 36]
    else:
        headers = ["编号", "标题", "优先级", "前置条件", "操作步骤", "预期结果", "接口校验点", "数据/UI 校验点"]
        widths = [16, 28, 8, 26, 40, 30, 30, 30]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, len(headers))

    for r, tc in enumerate(bundle.test_cases, 2):
        preconditions = "\n".join(f"• {p}" for p in tc.preconditions) if tc.preconditions else "-"
        steps = "\n".join(
            f"{s.seq}. {s.description}" for s in tc.steps
        ) if tc.steps else "-"

        # expected
        expected_parts = []
        resp = tc.expected.response
        if resp:
            status = resp.get("status", "")
            body_desc = ", ".join(f"{k}={v}" for k, v in resp.get("body", {}).items())
            if status or body_desc:
                expected_parts.append(f"HTTP {status}" if status else "")
                if body_desc:
                    expected_parts.append(body_desc)
        expected = "; ".join(expected_parts) if expected_parts else "-"

        # data assertions
        data_parts = []
        for da in tc.expected.data_assertions:
            data_parts.append(da.message or f"{da.target} {da.operator} {da.expected}")
        data_checks = "\n".join(f"• {d}" for d in data_parts) if data_parts else "-"

        # DOM assertions
        dom_parts = []
        for da in tc.expected.dom_assertions:
            dom_parts.append(da.message or f"{da.target} {da.operator} {da.expected}")

        if scope == Scope.BACKEND:
            ws.cell(row=r, column=1, value=tc.id)
            ws.cell(row=r, column=2, value=tc.title)
            ws.cell(row=r, column=3, value=tc.priority.value)
            ws.cell(row=r, column=4, value=preconditions)
            ws.cell(row=r, column=5, value=steps)
            ws.cell(row=r, column=6, value=expected)
            ws.cell(row=r, column=7, value=data_checks)

        elif scope == Scope.FRONTEND:
            ui_checks = "\n".join(f"• {d}" for d in dom_parts) if dom_parts else "-"
            ws.cell(row=r, column=1, value=tc.id)
            ws.cell(row=r, column=2, value=tc.title)
            ws.cell(row=r, column=3, value=tc.priority.value)
            ws.cell(row=r, column=4, value=preconditions)
            ws.cell(row=r, column=5, value=steps)
            ws.cell(row=r, column=6, value=expected if expected != "-" else ui_checks)
            ws.cell(row=r, column=7, value=ui_checks)

        else:  # fullstack
            http_checks = expected if expected != "-" else "-"
            dom_data_checks_parts = []
            for da in tc.expected.data_assertions:
                dom_data_checks_parts.append(da.message or f"{da.target} {da.operator} {da.expected}")
            for da in tc.expected.dom_assertions:
                dom_data_checks_parts.append(da.message or f"{da.target} {da.operator} {da.expected}")
            dom_data_checks = "\n".join(f"• {d}" for d in dom_data_checks_parts) if dom_data_checks_parts else "-"

            ws.cell(row=r, column=1, value=tc.id)
            ws.cell(row=r, column=2, value=tc.title)
            ws.cell(row=r, column=3, value=tc.priority.value)
            ws.cell(row=r, column=4, value=preconditions)
            ws.cell(row=r, column=5, value=steps)
            ws.cell(row=r, column=6, value=expected)
            ws.cell(row=r, column=7, value=http_checks)
            ws.cell(row=r, column=8, value=dom_data_checks)

    _style_data(ws, 2, len(bundle.test_cases) + 1, len(headers))
    _col_widths(ws, widths)
    ws.auto_filter.ref = ws.dimensions


def _write_sheet_coverage(wb: Workbook, bundle: TestDesignBundle) -> None:
    ws = wb.create_sheet("覆盖矩阵")
    headers = ["需求编号", "用例编号", "覆盖状态", "备注"]
    widths = [16, 36, 10, 30]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, len(headers))

    for r, cov in enumerate(bundle.coverage, 2):
        ws.cell(row=r, column=1, value=cov.requirement_ref)
        ws.cell(row=r, column=2, value=", ".join(cov.test_case_ids))
        ws.cell(row=r, column=3, value=cov.status.value)
        ws.cell(row=r, column=4, value=cov.notes or "-")

    _style_data(ws, 2, len(bundle.coverage) + 1, len(headers))
    _col_widths(ws, widths)
    ws.auto_filter.ref = ws.dimensions


def _write_sheet_open_questions(wb: Workbook, bundle: TestDesignBundle) -> None:
    if not bundle.open_questions:
        ws = wb.create_sheet("未决问题")
        ws.cell(row=1, column=1, value="无未决问题")
        return

    ws = wb.create_sheet("未决问题")
    headers = ["问题编号", "问题描述", "影响范围", "阻塞用例"]
    widths = [14, 44, 30, 36]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, len(headers))

    for r, q in enumerate(bundle.open_questions, 2):
        ws.cell(row=r, column=1, value=q.id)
        ws.cell(row=r, column=2, value=q.question)
        ws.cell(row=r, column=3, value=q.impact or "-")
        ws.cell(row=r, column=4, value=", ".join(q.blocks) if q.blocks else "-")

    _style_data(ws, 2, len(bundle.open_questions) + 1, len(headers))
    _col_widths(ws, widths)


def _write_summary_row(wb: Workbook, bundle: TestDesignBundle) -> None:
    ws = wb["测试用例"]
    total = len(bundle.test_cases)
    if total == 0:
        return

    normal = sum(1 for tc in bundle.test_cases
                 if any(t.value in ("functional",) for t in tc.test_types))
    negative = sum(1 for tc in bundle.test_cases
                   if any(t.value in ("negative",) for t in tc.test_types))
    boundary = sum(1 for tc in bundle.test_cases
                   if any(t.value in ("boundary",) for t in tc.test_types))
    permission = sum(1 for tc in bundle.test_cases
                     if any(t.value in ("permission",) for t in tc.test_types))

    p0p1_total = sum(1 for tc in bundle.test_cases if tc.priority.value in ("P0", "P1"))
    p0p1_covered = sum(
        1 for tc in bundle.test_cases
        if tc.priority.value in ("P0", "P1") and tc.requirement_refs
    )

    summary_row = total + 3
    text = (
        f"覆盖正常 {normal} 条 / 异常 {negative} 条 / 边界 {boundary} 条 / 权限 {permission} 条，"
        f"共 {total} 条。P0/P1 覆盖 {p0p1_covered}/{p0p1_total}。"
    )
    cell = ws.cell(row=summary_row, column=1, value=text)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=ws.max_column)
    cell.font = Font(bold=True, size=11, color="1F4E79")
